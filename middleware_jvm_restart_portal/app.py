import json
import os
import re
import sqlite3
import sys
from datetime import datetime, timezone
from functools import wraps
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlencode

BASE_DIR = Path(__file__).resolve().parent
VENDOR_DIR = BASE_DIR / "vendor" / "site-packages"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

import requests
from dotenv import load_dotenv
from flask import Flask, jsonify, redirect, render_template, request, session, url_for
from ldap3 import ALL, BASE, SUBTREE, Connection, Server
from ldap3.core.exceptions import LDAPException
from ldap3.utils.conv import escape_filter_chars
from openpyxl import load_workbook

load_dotenv(BASE_DIR / ".env")

DATA_DIR = BASE_DIR / "data"
DATA_FILE = DATA_DIR / "jvms.json"
EXCEL_INVENTORY_DEFAULT = DATA_DIR / "jvm_inventory.xlsx"
HISTORY_DB_DEFAULT = BASE_DIR / "history.sqlite3"
DEFAULT_ENVIRONMENTS = ["UNIT", "INTG", "PERF", "QUAL", "TRAINING", "PRODUCTION"]
PRODUCTION_ENVIRONMENTS = {"TRAINING", "PRODUCTION"}
TERMINAL_JOB_STATUSES = {"successful", "failed", "error", "canceled"}
EXCLUDED_INSTANCE_RE = re.compile(r"-i[1-9]$", re.IGNORECASE)


def str_to_bool(value: Optional[str], default: bool = False) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


class Settings:
    SECRET_KEY = os.getenv("APP_SECRET_KEY", "dev-only-change-me")
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = "Lax"
    SESSION_COOKIE_SECURE = str_to_bool(os.getenv("SESSION_COOKIE_SECURE"), False)
    PORTAL_PUBLIC_URL = os.getenv("PORTAL_PUBLIC_URL", "").rstrip("/")

    # Auth modes: local, ldap, saml_header
    # saml_header means SAML is handled before the app and the authenticated user is passed in a header.
    AUTH_MODE = os.getenv("AUTH_MODE", "").strip().lower()
    AUTH_REQUIRED = str_to_bool(os.getenv("AUTH_REQUIRED"), True)
    AUTH_DISABLED_USERNAME = os.getenv("AUTH_DISABLED_USERNAME", "local-dev-user")
    AUTH_DISABLED_DISPLAY_NAME = os.getenv("AUTH_DISABLED_DISPLAY_NAME", "Local Development User")
    SAML_ENABLED = str_to_bool(os.getenv("SAML_ENABLED"), False)
    SAML_LOGIN_URL = os.getenv("SAML_LOGIN_URL", "")
    SAML_LOGOUT_URL = os.getenv("SAML_LOGOUT_URL", "")
    SAML_REDIRECT_PARAM = os.getenv("SAML_REDIRECT_PARAM", "redirect")
    SAML_USER_HEADER = os.getenv("SAML_USER_HEADER", "X-Remote-User")
    SAML_DISPLAY_NAME_HEADER = os.getenv("SAML_DISPLAY_NAME_HEADER", "X-Remote-Display-Name")
    SAML_EMAIL_HEADER = os.getenv("SAML_EMAIL_HEADER", "X-Remote-Email")
    SAML_ID_MAP = os.getenv("SAML_ID_MAP", "local_realm")
    SAML_UNIQUE_ID_ATTRIBUTE = os.getenv("SAML_UNIQUE_ID_ATTRIBUTE", "UID")
    SAML_PRINCIPAL_NAME_ATTRIBUTE = os.getenv("SAML_PRINCIPAL_NAME_ATTRIBUTE", "UID")

    CONTROLLER_API_BASE = os.getenv("CONTROLLER_API_BASE", "").rstrip("/")
    CONTROLLER_TOKEN = os.getenv("CONTROLLER_TOKEN", "")
    CONTROLLER_VERIFY_SSL = str_to_bool(os.getenv("CONTROLLER_VERIFY_SSL"), True)
    JOB_TEMPLATE_ID = os.getenv("JOB_TEMPLATE_ID", "")

    PROD_CONTROLLER_API_BASE = os.getenv("PROD_CONTROLLER_API_BASE", "").rstrip("/")
    PROD_CONTROLLER_TOKEN = os.getenv("PROD_CONTROLLER_TOKEN", "")
    PROD_CONTROLLER_VERIFY_SSL = str_to_bool(os.getenv("PROD_CONTROLLER_VERIFY_SSL"), True)
    PROD_JOB_TEMPLATE_ID = os.getenv("PROD_JOB_TEMPLATE_ID", "")
    AUTO_LIMIT_FROM_JVMS = str_to_bool(os.getenv("AUTO_LIMIT_FROM_JVMS"), False)
    PAYLOAD_VAR_NAME = os.getenv("PAYLOAD_VAR_NAME", "jvm_restart")
    RITM_NUMBER = os.getenv("RITM_NUMBER", "RITMTEST")
    PORTAL_REQUESTED_BY = os.getenv("PORTAL_REQUESTED_BY", "middleware-portal")
    PROD_ENVIRONMENTS_ENABLED = str_to_bool(os.getenv("PROD_ENVIRONMENTS_ENABLED"), False)

    JVM_SOURCE_MODE = os.getenv("JVM_SOURCE_MODE", "excel").strip().lower()
    JVM_INVENTORY_FILE = os.getenv("JVM_INVENTORY_FILE", str(EXCEL_INVENTORY_DEFAULT))
    JVM_RESOURCE_URL = os.getenv("JVM_RESOURCE_URL", "")
    JVM_RESOURCE_TOKEN = os.getenv("JVM_RESOURCE_TOKEN", "")
    JVM_RESOURCE_VERIFY_SSL = str_to_bool(os.getenv("JVM_RESOURCE_VERIFY_SSL"), True)

    PROD_JVM_RESOURCE_URL = os.getenv("PROD_JVM_RESOURCE_URL", "")
    PROD_JVM_RESOURCE_TOKEN = os.getenv("PROD_JVM_RESOURCE_TOKEN", "")
    PROD_JVM_RESOURCE_VERIFY_SSL = str_to_bool(os.getenv("PROD_JVM_RESOURCE_VERIFY_SSL"), True)

    HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", "30"))
    HISTORY_DB_PATH = os.getenv("HISTORY_DB_PATH", str(HISTORY_DB_DEFAULT))
    HISTORY_LIMIT = int(os.getenv("HISTORY_LIMIT", "50"))

    LDAP_ENABLED = str_to_bool(os.getenv("LDAP_ENABLED"), False)
    LDAP_SERVER_URI = os.getenv("LDAP_SERVER_URI", "ldaps://ldap.example.com:636")
    LDAP_REQUIRE_SSL = str_to_bool(os.getenv("LDAP_REQUIRE_SSL"), True)
    LDAP_BIND_DN = os.getenv("LDAP_BIND_DN", "")
    LDAP_BIND_PASSWORD = os.getenv("LDAP_BIND_PASSWORD", "")
    LDAP_USER_BASE_DN = os.getenv("LDAP_USER_BASE_DN", "")
    LDAP_USER_FILTER = os.getenv("LDAP_USER_FILTER", "(uid={username})")
    LDAP_REQUIRED_GROUP_DN = os.getenv("LDAP_REQUIRED_GROUP_DN", "")
    LDAP_GROUP_MEMBER_ATTRIBUTE = os.getenv("LDAP_GROUP_MEMBER_ATTRIBUTE", "member")
    LDAP_DISPLAY_NAME_ATTRIBUTE = os.getenv("LDAP_DISPLAY_NAME_ATTRIBUTE", "cn")
    LDAP_UID_ATTRIBUTE = os.getenv("LDAP_UID_ATTRIBUTE", "uid")


app = Flask(__name__)
app.config.from_object(Settings)


class PortalError(Exception):
    def __init__(self, message: str, status_code: int = 400, details: Any = None):
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.details = details


@app.errorhandler(PortalError)
def handle_portal_error(exc: PortalError):
    payload = {"error": exc.message}
    if exc.details is not None:
        payload["details"] = exc.details
    return jsonify(payload), exc.status_code


@app.errorhandler(Exception)
def handle_unexpected_error(exc: Exception):
    return jsonify({"error": "Unexpected portal error", "details": str(exc)}), 500


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_environment(value: Any) -> str:
    return str(value or "").strip().upper()


def is_production_environment(env: str) -> bool:
    return normalize_environment(env) in PRODUCTION_ENVIRONMENTS


def is_excluded_instance_name(value: Any) -> bool:
    return bool(EXCLUDED_INSTANCE_RE.search(str(value or "").strip()))


def should_ingest_jvm(jvm_name: Any) -> bool:
    """Exclude JVM names ending in -i1 through -i9 before they reach the UI or payload."""
    return not is_excluded_instance_name(jvm_name)


def filter_ingested_jvms(jvms: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return [jvm for jvm in jvms if should_ingest_jvm(jvm.get("name", ""))]


def environment_is_visible(env: str) -> bool:
    normalized_env = normalize_environment(env)
    if is_production_environment(normalized_env) and not app.config["PROD_ENVIRONMENTS_ENABLED"]:
        return False
    return True


def resolve_project_path(path_value: str) -> Path:
    path = Path(path_value)
    if path.is_absolute():
        return path
    return BASE_DIR / path


def get_db_connection() -> sqlite3.Connection:
    db_path = resolve_project_path(app.config["HISTORY_DB_PATH"])
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_history_db() -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS job_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                username TEXT NOT NULL,
                job_id INTEGER,
                status TEXT NOT NULL,
                selected_items_json TEXT NOT NULL,
                launch_payload_json TEXT NOT NULL,
                controller_response_json TEXT,
                failure_message TEXT,
                controller_profile TEXT DEFAULT 'nonprod'
            )
            """
        )
        existing_columns = {row[1] for row in conn.execute("PRAGMA table_info(job_history)").fetchall()}
        if "controller_profile" not in existing_columns:
            conn.execute("ALTER TABLE job_history ADD COLUMN controller_profile TEXT DEFAULT 'nonprod'")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_job_history_job_id ON job_history(job_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_job_history_created_at ON job_history(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_job_history_controller_profile ON job_history(controller_profile)")


def record_job_history(
    username: str,
    items: List[Dict[str, str]],
    launch_payload: Dict[str, Any],
    job_id: Optional[int],
    controller_response: Optional[Dict[str, Any]],
    status: str = "launched",
    failure_message: Optional[str] = None,
    controller_profile: str = "nonprod",
) -> None:
    timestamp = now_utc()
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO job_history (
                created_at, updated_at, username, job_id, status,
                selected_items_json, launch_payload_json, controller_response_json, failure_message, controller_profile
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                timestamp,
                timestamp,
                username,
                job_id,
                status,
                json.dumps(items),
                json.dumps(launch_payload),
                json.dumps(controller_response or {}),
                failure_message,
                controller_profile,
            ),
        )


def update_job_history(job_id: int, status: str, failure_message: Optional[str] = None) -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            UPDATE job_history
            SET updated_at = ?, status = ?, failure_message = COALESCE(?, failure_message)
            WHERE job_id = ?
            """,
            (now_utc(), status, failure_message, job_id),
        )


def history_row_to_dict(row: sqlite3.Row) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "username": row["username"],
        "job_id": row["job_id"],
        "status": row["status"],
        "selected_items": json.loads(row["selected_items_json"] or "[]"),
        "launch_payload": json.loads(row["launch_payload_json"] or "{}"),
        "failure_message": row["failure_message"],
        "controller_profile": row["controller_profile"] if "controller_profile" in row.keys() else "nonprod",
    }


def list_job_history(limit: int) -> List[Dict[str, Any]]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, created_at, updated_at, username, job_id, status,
                   selected_items_json, launch_payload_json, failure_message, controller_profile
            FROM job_history
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [history_row_to_dict(row) for row in rows]


def list_active_job_history(limit: int = 25) -> List[Dict[str, Any]]:
    placeholders = ",".join(["?"] * len(TERMINAL_JOB_STATUSES))
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id, created_at, updated_at, username, job_id, status,
                   selected_items_json, launch_payload_json, failure_message, controller_profile
            FROM job_history
            WHERE lower(status) NOT IN ({placeholders})
            ORDER BY id DESC
            LIMIT ?
            """,
            tuple(TERMINAL_JOB_STATUSES) + (limit,),
        ).fetchall()
    return [history_row_to_dict(row) for row in rows]


def controller_profile_from_items(items: List[Dict[str, str]]) -> str:
    has_prod = any(is_production_environment(item.get("environment", "")) for item in items)
    has_nonprod = any(not is_production_environment(item.get("environment", "")) for item in items)
    if has_prod and has_nonprod:
        raise PortalError("Do not mix production/training JVMs with non-production JVMs in the same restart request", 400)
    return "prod" if has_prod else "nonprod"


def controller_config(profile: str) -> Dict[str, Any]:
    normalized_profile = (profile or "nonprod").strip().lower()
    if normalized_profile == "prod":
        return {
            "profile": "prod",
            "api_base": app.config["PROD_CONTROLLER_API_BASE"],
            "token": app.config["PROD_CONTROLLER_TOKEN"],
            "verify_ssl": app.config["PROD_CONTROLLER_VERIFY_SSL"],
            "job_template_id": app.config["PROD_JOB_TEMPLATE_ID"],
        }
    return {
        "profile": "nonprod",
        "api_base": app.config["CONTROLLER_API_BASE"],
        "token": app.config["CONTROLLER_TOKEN"],
        "verify_ssl": app.config["CONTROLLER_VERIFY_SSL"],
        "job_template_id": app.config["JOB_TEMPLATE_ID"],
    }


def controller_headers(profile: str = "nonprod", accept: str = "application/json", content_type: Optional[str] = "application/json") -> Dict[str, str]:
    cfg = controller_config(profile)
    token = cfg["token"]
    if not token:
        raise PortalError(f"{cfg['profile'].upper()} controller token is not configured", 500)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": accept,
    }
    if content_type:
        headers["Content-Type"] = content_type
    return headers


def controller_url(path: str, profile: str = "nonprod") -> str:
    cfg = controller_config(profile)
    base = cfg["api_base"]
    if not base:
        raise PortalError(f"{cfg['profile'].upper()} controller API base URL is not configured", 500)
    return f"{base}/{path.lstrip('/')}"


def controller_request(
    method: str,
    path: str,
    profile: str = "nonprod",
    accept: str = "application/json",
    content_type: Optional[str] = "application/json",
    **kwargs,
) -> requests.Response:
    cfg = controller_config(profile)
    response = requests.request(
        method=method,
        url=controller_url(path, profile),
        headers=controller_headers(profile, accept=accept, content_type=content_type),
        timeout=app.config["HTTP_TIMEOUT"],
        verify=cfg["verify_ssl"],
        **kwargs,
    )
    if response.status_code >= 400:
        try:
            details = response.json()
        except ValueError:
            details = response.text
        raise PortalError(
            f"{cfg['profile'].upper()} Controller API returned HTTP {response.status_code}",
            status_code=502,
            details=details,
        )
    return response


def get_controller_profile_for_job(job_id: int) -> str:
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT controller_profile FROM job_history WHERE job_id = ? ORDER BY id DESC LIMIT 1",
            (job_id,),
        ).fetchone()
    if row and row["controller_profile"]:
        return row["controller_profile"]
    return "nonprod"


def refresh_active_jobs_from_controller() -> None:
    for active_job in list_active_job_history(limit=25):
        job_id = active_job.get("job_id")
        if not job_id:
            continue
        try:
            response = controller_request("GET", f"jobs/{job_id}/", profile=get_controller_profile_for_job(job_id))
            job = response.json()
            status = job.get("status") or active_job.get("status") or "unknown"
            failure_message = job.get("job_explanation") or job.get("result_traceback")
            update_job_history(
                int(job_id),
                status,
                failure_message if status in TERMINAL_JOB_STATUSES else None,
            )
        except Exception:
            continue


def current_auth_mode() -> str:
    configured_mode = (app.config["AUTH_MODE"] or "").strip().lower()

    # Safety behavior:
    # SAML_ENABLED=false must disable SSO even if AUTH_MODE was accidentally left as saml_header.
    # This prevents the portal from continuing to redirect to SSO after SSO is turned off.
    if configured_mode == "saml_header" and not app.config["SAML_ENABLED"]:
        if app.config["LDAP_ENABLED"]:
            return "ldap"
        return "local"

    if configured_mode:
        return configured_mode
    if app.config["SAML_ENABLED"]:
        return "saml_header"
    if app.config["LDAP_ENABLED"]:
        return "ldap"
    return "local"


def external_url_for(path: str) -> str:
    public_url = app.config["PORTAL_PUBLIC_URL"]
    if public_url:
        return f"{public_url}/{path.lstrip('/')}"
    return url_for(path.strip("/") or "index", _external=True)


def build_saml_login_redirect(next_path: str) -> str:
    login_url = app.config["SAML_LOGIN_URL"]
    if not login_url:
        raise PortalError("SAML_LOGIN_URL is not configured", 500)

    callback_path = url_for("saml_callback", next=next_path)
    if app.config["PORTAL_PUBLIC_URL"]:
        callback_url = f"{app.config['PORTAL_PUBLIC_URL']}{callback_path}"
    else:
        callback_url = url_for("saml_callback", next=next_path, _external=True)

    separator = "&" if "?" in login_url else "?"
    redirect_param = app.config["SAML_REDIRECT_PARAM"] or "redirect"
    return f"{login_url}{separator}{urlencode({redirect_param: callback_url})}"


def get_header_value(header_name: str) -> str:
    if not header_name:
        return ""
    if header_name.upper() == "REMOTE_USER":
        return str(request.environ.get("REMOTE_USER") or "").strip()
    return str(request.headers.get(header_name) or request.environ.get(f"HTTP_{header_name.upper().replace('-', '_')}") or "").strip()


def get_saml_username_from_request() -> str:
    return get_header_value(app.config["SAML_USER_HEADER"])


def authenticate_local(username: str, password: str) -> Tuple[bool, str, str]:
    if username and password:
        return True, username, ""
    return False, "", "Username and password are required"


def find_and_authorize_ldap_user(username: str) -> Tuple[bool, str, str, str]:
    if not username:
        return False, "", "Username is required", ""
    if not app.config["LDAP_USER_BASE_DN"]:
        return False, "", "LDAP_USER_BASE_DN is not configured", ""
    if not app.config["LDAP_BIND_DN"] or not app.config["LDAP_BIND_PASSWORD"]:
        return False, "", "LDAP service bind is not configured", ""
    if app.config["LDAP_REQUIRE_SSL"] and not app.config["LDAP_SERVER_URI"].lower().startswith("ldaps://"):
        return False, "", "LDAP_REQUIRE_SSL is true but LDAP_SERVER_URI does not use ldaps://", ""

    safe_username = escape_filter_chars(username)
    user_filter = app.config["LDAP_USER_FILTER"].format(username=safe_username)

    try:
        server = Server(app.config["LDAP_SERVER_URI"], get_info=ALL)
        service_conn = Connection(
            server,
            user=app.config["LDAP_BIND_DN"],
            password=app.config["LDAP_BIND_PASSWORD"],
            auto_bind=True,
        )
        service_conn.search(
            search_base=app.config["LDAP_USER_BASE_DN"],
            search_filter=user_filter,
            search_scope=SUBTREE,
            attributes=[
                app.config["LDAP_DISPLAY_NAME_ATTRIBUTE"],
                "memberOf",
                "distinguishedName",
                "userPrincipalName",
                "sAMAccountName",
                app.config["LDAP_UID_ATTRIBUTE"],
            ],
            size_limit=1,
        )
        if not service_conn.entries:
            service_conn.unbind()
            return False, "", "User was not found in LDAP", ""

        user_entry = service_conn.entries[0]
        user_dn = user_entry.entry_dn
        display_attr = app.config["LDAP_DISPLAY_NAME_ATTRIBUTE"]
        display_name = str(user_entry[display_attr].value) if display_attr in user_entry else username

        required_group_dn = app.config["LDAP_REQUIRED_GROUP_DN"]
        if required_group_dn:
            member_of_values = []
            if "memberOf" in user_entry:
                raw_member_of = user_entry["memberOf"].values or []
                member_of_values = [str(value).lower() for value in raw_member_of]

            direct_member = required_group_dn.lower() in member_of_values
            if not direct_member:
                group_filter = f"({app.config['LDAP_GROUP_MEMBER_ATTRIBUTE']}={escape_filter_chars(user_dn)})"
                service_conn.search(
                    search_base=required_group_dn,
                    search_filter=group_filter,
                    search_scope=BASE,
                    attributes=[app.config["LDAP_GROUP_MEMBER_ATTRIBUTE"]],
                    size_limit=1,
                )
                direct_member = bool(service_conn.entries)

            if not direct_member:
                service_conn.unbind()
                return False, "", "User is not authorized for the required LDAP group", user_dn

        service_conn.unbind()
        return True, display_name, "", user_dn

    except LDAPException as exc:
        return False, "", f"LDAP authorization failed: {exc}", ""


def authenticate_ldap(username: str, password: str) -> Tuple[bool, str, str]:
    if not username or not password:
        return False, "", "Username and password are required"

    ok, display_name, error, user_dn = find_and_authorize_ldap_user(username)
    if not ok:
        return False, "", error

    try:
        server = Server(app.config["LDAP_SERVER_URI"], get_info=ALL)
        user_conn = Connection(server, user=user_dn, password=password, auto_bind=True)
        user_conn.unbind()
        return True, display_name, ""
    except LDAPException as exc:
        return False, "", f"LDAP authentication failed: {exc}"


def authenticate_saml_header() -> Tuple[bool, str, str, str]:
    username = get_saml_username_from_request()
    if not username:
        return False, "", "SAML user header was not found", ""

    ok, display_name, error, _ = find_and_authorize_ldap_user(username)
    if not ok:
        return False, username, error, ""

    header_display_name = get_header_value(app.config["SAML_DISPLAY_NAME_HEADER"])
    return True, username, header_display_name or display_name or username, ""


def authenticate_user(username: str, password: str) -> Tuple[bool, str, str]:
    mode = current_auth_mode()
    if mode == "ldap":
        return authenticate_ldap(username, password)
    if mode == "saml_header":
        return False, "", "SAML users must sign in through the configured SAML redirect URL"
    return authenticate_local(username, password)


def create_session(username: str, display_name: str) -> None:
    session.clear()
    session["username"] = username
    session["display_name"] = display_name or username


def create_session_when_auth_disabled() -> None:
    if not session.get("username"):
        create_session(app.config["AUTH_DISABLED_USERNAME"], app.config["AUTH_DISABLED_DISPLAY_NAME"])


def create_session_from_saml_header() -> Tuple[bool, str]:
    ok, username, display_name, error = authenticate_saml_header()
    if ok:
        create_session(username, display_name)
        return True, ""
    return False, error


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not app.config["AUTH_REQUIRED"]:
            create_session_when_auth_disabled()
            return view_func(*args, **kwargs)

        if session.get("username"):
            return view_func(*args, **kwargs)

        if current_auth_mode() == "saml_header":
            ok, error = create_session_from_saml_header()
            if ok:
                return view_func(*args, **kwargs)

        if request.path.startswith("/api/"):
            payload = {"error": "Authentication required"}
            if current_auth_mode() == "saml_header" and app.config["SAML_LOGIN_URL"]:
                payload["login_url"] = build_saml_login_redirect(request.path)
            return jsonify(payload), 401
        return redirect(url_for("login", next=request.path))

    return wrapped


def normalize_jvm_entry(item: Any) -> Dict[str, str]:
    if isinstance(item, str):
        return {"name": item, "display_name": item, "host": ""}
    if not isinstance(item, dict):
        raise PortalError("Invalid JVM inventory item", 500, item)

    name = item.get("name") or item.get("jvm_name") or item.get("server")
    if not name:
        raise PortalError("JVM item is missing name/jvm_name/server", 500, item)

    return {
        "name": str(name),
        "display_name": str(item.get("display_name") or item.get("label") or name),
        "host": str(item.get("host") or item.get("hostname") or ""),
    }


def load_excel_inventory() -> Dict[str, List[Dict[str, str]]]:
    inventory_file = resolve_project_path(app.config["JVM_INVENTORY_FILE"])
    if not inventory_file.exists():
        raise PortalError(f"JVM inventory Excel file not found: {inventory_file}", 500)

    workbook = load_workbook(inventory_file, read_only=True, data_only=True)
    try:
        worksheet = workbook.active

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise PortalError("JVM inventory Excel sheet is empty", 500)

        headers = {str(value).strip().upper(): index for index, value in enumerate(header_row) if value is not None}
        required_columns = ["ENVIRONMENT", "JVM_NAME", "UNIX_HOST"]
        missing_columns = [column for column in required_columns if column not in headers]
        if missing_columns:
            raise PortalError(
                "JVM inventory Excel file is missing required columns",
                500,
                {"missing_columns": missing_columns, "required_columns": required_columns},
            )

        inventory: Dict[str, List[Dict[str, str]]] = {}
        seen = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            env = normalize_environment(row[headers["ENVIRONMENT"]])
            jvm_name = str(row[headers["JVM_NAME"]] or "").strip()
            unix_host = str(row[headers["UNIX_HOST"]] or "").strip()

            if not env and not jvm_name and not unix_host:
                continue
            if not env or not jvm_name or not unix_host:
                continue
            if not should_ingest_jvm(jvm_name):
                continue

            key = (env, unix_host, jvm_name)
            if key in seen:
                continue
            seen.add(key)

            inventory.setdefault(env, []).append(
                {"name": jvm_name, "display_name": jvm_name, "host": unix_host}
            )

        for env in inventory:
            inventory[env] = sorted(inventory[env], key=lambda item: (item["name"], item["host"]))
        return dict(sorted(inventory.items(), key=lambda item: item[0]))
    finally:
        workbook.close()


def extract_jvms_from_api_payload(payload: Any) -> List[Dict[str, str]]:
    if isinstance(payload, list):
        raw_items = payload
    elif isinstance(payload, dict):
        raw_items = payload.get("jvms") or payload.get("results") or payload.get("items") or []
    else:
        raw_items = []
    return sorted(filter_ingested_jvms([normalize_jvm_entry(item) for item in raw_items]), key=lambda x: x["name"])


def load_json_jvms(env: str) -> List[Dict[str, str]]:
    if not DATA_FILE.exists():
        raise PortalError(f"JVM JSON file not found: {DATA_FILE}", 500)
    with DATA_FILE.open("r", encoding="utf-8") as file_handle:
        data = json.load(file_handle)
    return sorted(filter_ingested_jvms([normalize_jvm_entry(item) for item in data.get(env.lower(), [])]), key=lambda x: x["name"])


def load_api_jvms(env: str) -> List[Dict[str, str]]:
    normalized_env = normalize_environment(env)
    use_prod_resource = is_production_environment(normalized_env) and bool(app.config["PROD_JVM_RESOURCE_URL"])
    resource_url = app.config["PROD_JVM_RESOURCE_URL"] if use_prod_resource else app.config["JVM_RESOURCE_URL"]
    if not resource_url:
        raise PortalError("JVM_RESOURCE_URL is required when JVM_SOURCE_MODE=api", 500)

    headers = {"Accept": "application/json"}
    token = app.config["PROD_JVM_RESOURCE_TOKEN"] if use_prod_resource else app.config["JVM_RESOURCE_TOKEN"]
    if token:
        headers["Authorization"] = f"Bearer {token}"

    response = requests.get(
        resource_url,
        params={"env": normalized_env},
        headers=headers,
        timeout=app.config["HTTP_TIMEOUT"],
        verify=app.config["PROD_JVM_RESOURCE_VERIFY_SSL"] if use_prod_resource else app.config["JVM_RESOURCE_VERIFY_SSL"],
    )
    if response.status_code >= 400:
        raise PortalError(f"JVM resource API returned HTTP {response.status_code}", 502, response.text)
    return extract_jvms_from_api_payload(response.json())


def load_environment_names() -> List[str]:
    if app.config["JVM_SOURCE_MODE"] == "excel":
        environments = list(load_excel_inventory().keys())
    else:
        environments = DEFAULT_ENVIRONMENTS
    return [env for env in environments if environment_is_visible(env)]


def load_jvms(env: str) -> List[Dict[str, str]]:
    normalized_env = normalize_environment(env)
    if not environment_is_visible(normalized_env):
        raise PortalError(f"Environment '{normalized_env}' is not visible because production environments are disabled", 403)
    if app.config["JVM_SOURCE_MODE"] == "excel":
        inventory = load_excel_inventory()
        if normalized_env not in inventory:
            raise PortalError(f"Invalid environment: {env}", 400)
        return inventory[normalized_env]
    if app.config["JVM_SOURCE_MODE"] == "api":
        return load_api_jvms(normalized_env)
    if normalized_env not in DEFAULT_ENVIRONMENTS:
        raise PortalError(f"Invalid environment: {env}", 400)
    return load_json_jvms(normalized_env)


def validate_restart_items(items: Any) -> Tuple[List[Dict[str, str]], Dict[str, Dict[str, Dict[str, str]]]]:
    if not isinstance(items, list) or not items:
        raise PortalError("At least one JVM must be selected before restart", 400)

    validated: List[Dict[str, str]] = []
    seen = set()
    env_cache: Dict[str, Dict[str, Dict[str, str]]] = {}

    for item in items:
        env = normalize_environment(item.get("environment") or item.get("env") or "")
        jvm_name = str(item.get("jvm_name") or item.get("name") or "").strip()
        if not env:
            raise PortalError("Invalid environment selected", 400)
        if not jvm_name:
            raise PortalError("JVM name is required", 400)

        if env not in env_cache:
            env_cache[env] = {jvm["name"]: jvm for jvm in load_jvms(env)}
        if jvm_name not in env_cache[env]:
            raise PortalError(f"JVM '{jvm_name}' was not found in environment '{env}'", 400)

        jvm_record = env_cache[env][jvm_name]
        hostname = str(item.get("host") or item.get("hostname") or jvm_record.get("host") or jvm_name).strip()

        key = (env, hostname, jvm_name)
        if key in seen:
            continue
        seen.add(key)
        validated.append({"environment": env, "host": hostname, "jvm_name": jvm_name})

    return validated, env_cache


def build_jvm_restart_payload(items: List[Dict[str, str]]) -> Dict[str, Any]:
    restart_data: Dict[str, Any] = {"ritm_number": app.config["RITM_NUMBER"], "envs": {}}
    for item in items:
        # UI/inventory uses normalized uppercase env names internally, but the Ansible payload expects lowercase keys.
        env = normalize_environment(item["environment"]).lower()
        hostname = item.get("host") or item["jvm_name"]
        jvm_name = item["jvm_name"]
        restart_data["envs"].setdefault(env, {"hosts": {}})
        restart_data["envs"][env]["hosts"][hostname] = jvm_name
    return {app.config["PAYLOAD_VAR_NAME"]: restart_data}


@app.route("/login", methods=["GET", "POST"])
def login():
    if not app.config["AUTH_REQUIRED"]:
        create_session_when_auth_disabled()
        return redirect(request.args.get("next") or url_for("index"))

    if session.get("username"):
        return redirect(url_for("index"))

    next_path = request.args.get("next") or url_for("index")
    mode = current_auth_mode()

    if mode == "saml_header":
        ok, error = create_session_from_saml_header()
        if ok:
            return redirect(next_path)
        if app.config["SAML_LOGIN_URL"]:
            return redirect(build_saml_login_redirect(next_path))
        return render_template("login.html", error=error or "SAML login is not configured", ldap_enabled=True, auth_mode=mode)

    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        ok, display_name, error = authenticate_user(username, password)
        if ok:
            create_session(username, display_name or username)
            return redirect(next_path)
    return render_template("login.html", error=error, ldap_enabled=app.config["LDAP_ENABLED"], auth_mode=mode)


@app.route("/saml/callback")
def saml_callback():
    next_path = request.args.get("next") or url_for("index")
    ok, error = create_session_from_saml_header()
    if ok:
        return redirect(next_path)
    return render_template("login.html", error=error or "SAML login failed LDAP authorization", ldap_enabled=True, auth_mode="saml_header")


@app.route("/logout")
def logout():
    session.clear()
    if current_auth_mode() == "saml_header" and app.config["SAML_LOGOUT_URL"]:
        return redirect(app.config["SAML_LOGOUT_URL"])
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return render_template("index.html", username=session.get("username"), display_name=session.get("display_name") or session.get("username"))


@app.route("/health")
def health():
    return jsonify({"status": "ok", "service": "middleware-jvm-restart"})


@app.route("/api/session")
@login_required
def api_session():
    return jsonify({"username": session.get("username"), "display_name": session.get("display_name")})


@app.route("/api/environments")
@login_required
def api_environments():
    return jsonify({"environments": load_environment_names()})


@app.route("/api/jvms")
@login_required
def api_jvms():
    env = request.args.get("env", "").strip()
    return jsonify({"environment": normalize_environment(env), "jvms": load_jvms(env)})


@app.route("/api/history")
@login_required
def api_history():
    limit = min(int(request.args.get("limit", app.config["HISTORY_LIMIT"])), 200)
    return jsonify({"history": list_job_history(limit)})


@app.route("/api/active-jobs")
@login_required
def api_active_jobs():
    refresh_active_jobs_from_controller()
    return jsonify({"active_jobs": list_active_job_history(limit=25)})


@app.route("/api/restart", methods=["POST"])
@login_required
def api_restart():
    body = request.get_json(force=True, silent=False)
    items, _ = validate_restart_items(body.get("items"))
    portal_payload = build_jvm_restart_payload(items)
    launch_payload: Dict[str, Any] = {"extra_vars": portal_payload}
    if app.config["AUTO_LIMIT_FROM_JVMS"]:
        launch_payload["limit"] = ",".join([item["jvm_name"] for item in items])

    controller_profile = controller_profile_from_items(items)
    cfg = controller_config(controller_profile)
    job_template_id = cfg["job_template_id"]
    if not job_template_id:
        record_job_history(
            session["username"], items, launch_payload, None, None, "failed",
            f"{controller_profile.upper()} job template ID is not configured",
            controller_profile=controller_profile,
        )
        raise PortalError(f"{controller_profile.upper()} job template ID is not configured", 500)

    response = controller_request("POST", f"job_templates/{job_template_id}/launch/", profile=controller_profile, data=json.dumps(launch_payload))
    data = response.json()
    job_id = data.get("job") or data.get("id")
    if not job_id:
        record_job_history(
            session["username"], items, launch_payload, None, data, "failed",
            "Controller launch response did not include job id",
            controller_profile=controller_profile,
        )
        raise PortalError("Controller launch response did not include job id", 502, data)

    record_job_history(session["username"], items, launch_payload, int(job_id), data, "launched", controller_profile=controller_profile)
    return jsonify({
        "message": "Restart job launched",
        "job_id": job_id,
        "controller_profile": controller_profile,
        "launch_payload": launch_payload,
        "controller_response": data,
    })


@app.route("/api/jobs/<int:job_id>/status")
@login_required
def api_job_status(job_id: int):
    response = controller_request("GET", f"jobs/{job_id}/", profile=get_controller_profile_for_job(job_id))
    job = response.json()
    status = job.get("status") or "unknown"
    failure_message = job.get("job_explanation") or job.get("result_traceback")
    update_job_history(job_id, status, failure_message if status in TERMINAL_JOB_STATUSES else None)
    return jsonify({
        "id": job.get("id"),
        "name": job.get("name"),
        "status": status,
        "failed": job.get("failed"),
        "started": job.get("started"),
        "finished": job.get("finished"),
        "elapsed": job.get("elapsed"),
        "job_explanation": job.get("job_explanation"),
        "result_traceback": job.get("result_traceback"),
        "artifacts": job.get("artifacts") or {},
        "url": job.get("url"),
    })


@app.route("/api/jobs/<int:job_id>/stdout")
@login_required
def api_job_stdout(job_id: int):
    response = controller_request(
        "GET",
        f"jobs/{job_id}/stdout/?format=txt",
        profile=get_controller_profile_for_job(job_id),
        accept="text/plain, */*",
        content_type=None,
    )
    return jsonify({"job_id": job_id, "stdout": response.text})


@app.route("/api/jobs/<int:job_id>/artifacts")
@login_required
def api_job_artifacts(job_id: int):
    response = controller_request("GET", f"jobs/{job_id}/", profile=get_controller_profile_for_job(job_id))
    job = response.json()
    return jsonify({"job_id": job_id, "artifacts": job.get("artifacts") or {}})


init_history_db()


if __name__ == "__main__":
    app.run(host=os.getenv("FLASK_HOST", "0.0.0.0"), port=int(os.getenv("FLASK_PORT", "5000")), debug=str_to_bool(os.getenv("FLASK_DEBUG"), False))
